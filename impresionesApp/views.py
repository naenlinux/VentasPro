from django.http import HttpResponse
from django.template.loader import get_template
from weasyprint import HTML
from ventaApp.models import Ventas, PedidosDetalle
from empresaApp.models import Empresa
from inventarioApp.models import Almacen
from mantenedoresApp.models import Productos, UnidadMedida
from pprint import pprint
from openpyxl import Workbook
import io
from openpyxl.styles import colors, Font, Color, PatternFill, Border, Side
from datetime import date, datetime


def generar_pdf(request, idVenta):
    venta = Ventas.objects.select_related('idPedido','tipComprobante').get(pk=idVenta)
    empresa = Empresa.objects.first()
    detalles = PedidosDetalle.objects.filter(idPedido=venta.idPedido)
    importe_total = venta.idPedido.subtotal + venta.idPedido.igv_total

    context = {
        'venta': venta,
        'empresa': empresa,
        'detalles': detalles,
        'importe_total': importe_total
    }
    # Recupera la plantilla HTML
    template = get_template('comprobanteVenta.html')
    pprint(venta)
    html_string = template.render(context)

    # Convierte el HTML en un documento PDF
    html = HTML(string=html_string)
    html.page_size = 'A4'
    pdf_file = html.write_pdf()

    # Crea una respuesta HTTP con el PDF
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="comprobanteVenta.pdf"'

    return response

def export_to_excel(request):
    almacen = Almacen.objects.select_related('idProducto__unidadMedida').filter(activo=1)

    wb = Workbook()
    ws = wb.active
    fecha_hoy = date.today()
    fecha_datetime = datetime.strptime(str(fecha_hoy),"%Y-%m-%d")
    fecha_forma = fecha_datetime.strftime("%d/%m/%y")

   
    hora = datetime.now().time()
    objeto_tiempo = datetime.strptime(str(hora), "%H:%M:%S.%f")
    hora_form = objeto_tiempo.strftime("%H:%M")
    ws['B2'] = 'REPORTE DE ALMACEN ' + fecha_forma +' '+str(hora_form)
    ws['B2'].fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00',fill_type='solid')
    ws['B2'].font = Font(bold=True)

    ws.merge_cells('B2:D2')

    ws['B4'] = 'PRODUCTO'
    ws['C4'] = 'STOCK X U/M'
    ws['D4'] = 'STOCK X UNIDAD'


    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)
    ws['D4'].font = Font(bold=True)

    ws['B4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['C4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')
    ws['D4'].fill = PatternFill(start_color='00FF9900', end_color='00FF9900',fill_type='solid')

    cont = 5

    for almace in almacen:
        ws.cell(row = cont, column = 2).value = almace.idProducto.nombre
        ws.cell(row = cont, column = 3).value = str(almace.cantidad) +' '+almace.idProducto.unidadMedida.nombre
        ws.cell(row = cont, column = 4).value = almace.total
        cont += 1

    for row in ws.iter_rows(min_row=4, max_row=cont-1, min_col=2, max_col=4):
        for cell in row:
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = border
    
    ancho_columnas = ['B','C','D']
    ancho_personalizado = 20

    for columna in ancho_columnas:
        ws.column_dimensions[columna].width = ancho_personalizado

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ReporteAlmacenExcel.xlsx"'

    # Guarda el libro de trabajo en un objeto BytesIO en lugar de disco
    #excel_file = io.BytesIO()
    wb.save(response)
    #response.write(excel_file.getvalue())

    return response